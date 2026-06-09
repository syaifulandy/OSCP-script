import re
import csv
import json
from pathlib import Path
from urllib.parse import urlparse, urlunparse
from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ============================================================
# CONFIG
# ============================================================

OPEN_PORTS_FILE = "global_open_ports.txt"
FFUF_FILE = "global_ffuf.txt"
NUCLEI_FILES = [
    "global_nuclei.txt",
    "general_nuclei.txt",
    "global_nuclei_nonweb.txt",
]

OUTPUT_FILE = "Auto Enum Proactive Subdom Telkom.xlsx"
EXCEL_CELL_SAFE_LIMIT = 32000


# ============================================================
# HELPERS
# ============================================================

def clean_host(raw: str) -> str:
    raw = (raw or "").strip()
    if re.match(r"^[A-Za-z0-9_.-]+:\d+$", raw):
        return raw.rsplit(":", 1)[0]
    return raw


def normalize_url(url: str) -> str:
    url = (url or "").strip().replace("\\_", "_")
    try:
        parsed = urlparse(url)
        if not parsed.scheme or not parsed.netloc:
            return url

        path = re.sub(r"/+", "/", parsed.path or "/")
        return urlunparse(
            (
                parsed.scheme.lower(),
                parsed.netloc.lower(),
                path,
                "",
                parsed.query,
                "",
            )
        )
    except Exception:
        return url


def url_host_port(url: str):
    parsed = urlparse(url)
    host = (parsed.hostname or "").lower()
    port = parsed.port

    if port is None:
        if parsed.scheme.lower() == "https":
            port = 443
        elif parsed.scheme.lower() == "http":
            port = 80
        else:
            port = ""

    return host, str(port)


def target_host_port_url(target: str):
    target = (target or "").strip().replace("\\_", "_")

    if target.startswith(("http://", "https://")):
        url = normalize_url(target)
        host, port = url_host_port(url)
        return host, port, url

    if ":" in target:
        host, port = target.rsplit(":", 1)
        return clean_host(host).lower(), str(port).strip(), target.lower()

    return clean_host(target).lower(), "", target.lower()


def severity_normalize(severity: str) -> str:
    severity = (severity or "").strip().capitalize()
    if severity in ["Info", "Informational", "Unknown", ""]:
        return "Low"
    if severity not in ["Critical", "High", "Medium", "Low"]:
        return "Low"
    return severity


def nuclei_title(template_id: str) -> str:
    template_id = (template_id or "").strip()
    base = template_id.split(":", 1)[0]

    mapping = {
        "CVE-2023-48795": "SSH Terrapin Attack Vulnerability (CVE-2023-48795)",
        "ssh-weak-algo-supported": "SSH Weak Algorithms Supported",
        "open-redirect-generic": "Open Redirect",
        "laravel-debug-enabled": "Laravel Debug Enabled",
        "node-express-dev-env": "Node Express Development Environment",
        "exposed-redis": "Exposed Redis",
        "redis-default-logins": "Redis Default/Empty Login",
        "pgsql-empty-password": "PostgreSQL Empty Password",
        "pgsql-default-db": "PostgreSQL Default Database Login",
        "postgres-default-logins": "PostgreSQL Default Login",
        "exposed-zookeeper": "Exposed ZooKeeper",
        "ldap-anonymous-login-detect": "LDAP Anonymous Login Detected",
    }

    if base in mapping:
        return mapping[base]
    if base.upper().startswith("CVE-"):
        return base
    return base.replace("-", " ").replace("_", " ").title()


def safe_join(items, sep="\n", limit=EXCEL_CELL_SAFE_LIMIT):
    """
    Join parsed values safely for Excel cell size limit.
    This is parsing-only; no recommendation is generated.
    """
    items = [str(item) for item in items if str(item).strip()]
    if not items:
        return ""

    output = []
    current_length = 0

    for index, item in enumerate(items):
        extra_length = len(item) + (len(sep) if output else 0)
        if current_length + extra_length > limit:
            remaining = len(items) - index
            note = f"{sep}[TRUNCATED_FOR_EXCEL_CELL_LIMIT: {remaining} more item(s)]"
            if current_length + len(note) <= limit:
                output.append(note.strip())
            break
        output.append(item)
        current_length += extra_length

    return sep.join(output)


def format_ffuf_row(row: tuple) -> str:
    url, status, size, redirect, final = row
    details = []
    if status:
        details.append(f"status={status}")
    if size:
        details.append(f"size={size}")
    if redirect:
        details.append(f"redirect={redirect}")
    if final:
        details.append(f"final={final}")

    if details:
        return f"{url} [{' | '.join(details)}]"
    return url


def format_nuclei_finding(record: dict) -> str:
    parts = [
        f"[{record['Severity Risk']}]",
        record["Nama Temuan"],
        f"({record['Template ID']})" if record.get("Template ID") else "",
        "-",
        record.get("URL / Target", ""),
    ]
    line = " ".join([part for part in parts if part])

    if record.get("Evidence"):
        line += f" [{record['Evidence']}]"

    return line


# ============================================================
# PARSE OPEN PORTS
# ============================================================

ports = []
ports_seen = set()

open_ports_path = Path(OPEN_PORTS_FILE)
if open_ports_path.exists():
    with open_ports_path.open("r", encoding="utf-8", errors="ignore") as file:
        for line in file:
            line = line.strip()
            if not line:
                continue

            parts = line.split(";")
            if len(parts) < 4:
                continue

            raw_host, proto, port, service = parts[:4]
            banner = ";".join(parts[4:]).strip() if len(parts) > 4 else ""

            host = clean_host(raw_host).lower()
            port = str(port).strip()
            protocol = proto.strip().upper()
            service_value = service.strip()

            if banner and banner not in service_value:
                service_value = f"{service_value} ({banner})"

            dedup_key = (host, port, protocol, service_value)
            if dedup_key in ports_seen:
                continue

            ports_seen.add(dedup_key)
            ports.append(
                {
                    "Host / FQDN": host,
                    "IP Address": "",
                    "Port": port,
                    "Service": service_value,
                    "Protocol": protocol,
                }
            )
else:
    print(f"[WARN] File not found, skipped: {OPEN_PORTS_FILE}")


# ============================================================
# PARSE FFUF
# ============================================================

ffuf_by_host_port = defaultdict(list)
ffuf_seen = set()

ffuf_path = Path(FFUF_FILE)
if ffuf_path.exists():
    with ffuf_path.open("r", encoding="utf-8", errors="ignore", newline="") as file:
        reader = csv.reader(file)
        next(reader, None)  # skip header

        for fields in reader:
            if len(fields) < 5:
                continue

            url, status, size, redirect, final = [item.strip().replace("\\_", "_") for item in fields[:5]]
            if not url.startswith(("http://", "https://")):
                continue

            url = normalize_url(url)
            host, port = url_host_port(url)
            if not host or not port:
                continue

            row = (url, status, size, redirect, final)
            dedup_key = (host, port, url, status, size, redirect, final)
            if dedup_key in ffuf_seen:
                continue

            ffuf_seen.add(dedup_key)
            ffuf_by_host_port[(host, port)].append(row)
else:
    print(f"[WARN] File not found, skipped: {FFUF_FILE}")


# ============================================================
# PARSE NUCLEI
# ============================================================

def parse_nuclei_file(file_name: str):
    records = []
    path = Path(file_name)

    if not path.exists():
        print(f"[WARN] File not found, skipped: {file_name}")
        return records

    pattern = re.compile(
        r"^\[(.*?)\]\s+\[(.*?)\]\s+\[(.*?)\]\s+(.+?)(?:\s+\[(.*)\])?\s*$"
    )

    with path.open("r", encoding="utf-8", errors="ignore") as file:
        for line_number, line in enumerate(file, 1):
            raw_line = line.strip().rstrip(",")
            if not raw_line:
                continue

            match = pattern.match(raw_line)
            if not match:
                continue

            template_id, protocol, severity, target, evidence = match.groups()
            template_id = (template_id or "").strip()
            protocol = (protocol or "").strip()
            severity = severity_normalize(severity)
            target = (target or "").strip()
            evidence = (evidence or "").strip()
            host, port, url = target_host_port_url(target)

            records.append(
                {
                    "Source File": file_name,
                    "Line": line_number,
                    "Template ID": template_id,
                    "Nama Temuan": nuclei_title(template_id),
                    "Protocol": protocol,
                    "Severity Risk": severity,
                    "Host / FQDN": host,
                    "Port": port,
                    "URL / Target": url,
                    "Evidence": evidence,
                }
            )

    return records


nuclei_records = []
nuclei_by_host_port = defaultdict(list)
nuclei_seen = set()

for nuclei_file in NUCLEI_FILES:
    for record in parse_nuclei_file(nuclei_file):
        dedup_key = (
            record["Template ID"],
            record["Protocol"],
            record["Severity Risk"],
            record["Host / FQDN"],
            record["Port"],
            record["URL / Target"],
            record["Evidence"],
        )

        if dedup_key in nuclei_seen:
            continue

        nuclei_seen.add(dedup_key)
        nuclei_records.append(record)
        nuclei_by_host_port[(record["Host / FQDN"], record["Port"])].append(record)


# ============================================================
# ============================================================
# SHEET 1: RESULTS
# Host / FQDN dan IP Address digabung menjadi 1 kolom: Host / IP Address.
# Row Results hanya dibuat dari global_open_ports.txt agar port tidak berubah karena FFUF/Nuclei.
# ============================================================

results_rows = []
open_port_keys = {(row["Host / FQDN"], str(row["Port"])) for row in ports}

for port_row in ports:
    host = port_row["Host / FQDN"].lower()
    ip_address = (port_row.get("IP Address", "") or "").strip()
    host_ip_value = host if not ip_address else f"{host}\n{ip_address}"
    port = str(port_row["Port"])
    key = (host, port)

    ffuf_items = sorted({format_ffuf_row(row) for row in ffuf_by_host_port.get(key, [])})
    nuclei_items = sorted({format_nuclei_finding(item) for item in nuclei_by_host_port.get(key, [])})

    results_rows.append(
        {
            "No": len(results_rows) + 1,
            "Host / IP Address": host_ip_value,
            "Port": port,
            "Service": port_row.get("Service", ""),
            "Protocol": port_row.get("Protocol", ""),
            "FFUF URLs": safe_join(ffuf_items),
            "Nuclei Findings": safe_join(nuclei_items),
        }
    )

# Data FFUF/Nuclei yang host+port-nya tidak ada di global_open_ports.txt tidak dibuat row baru.
# Tujuannya agar port di Results tetap mengikuti file port scan.
unmatched_ffuf_keys = set(ffuf_by_host_port.keys()) - open_port_keys
unmatched_nuclei_keys = set(nuclei_by_host_port.keys()) - open_port_keys

results_df = pd.DataFrame(
    results_rows,
    columns=[
        "No",
        "Host / IP Address",
        "Port",
        "Service",
        "Protocol",
        "FFUF URLs",
        "Nuclei Findings",
    ],
)


# ============================================================
# SHEET 2: REMEDIASI - MANUAL ONLY / EMPTY
# ============================================================

remediation_df = pd.DataFrame(
    columns=[
        "No",
        "Category",
        "Finding Summary",
        "Recommendation",
        "Priority",
    ]
)


# ============================================================
# SHEET 3: CYSVM - KOLOM TETAP
# ============================================================

cysvm_rows = []
cysvm_grouped = defaultdict(list)

for record in nuclei_records:
    key = (
        record["Nama Temuan"],
        record["Severity Risk"],
        record["Host / FQDN"],
        record["Port"],
    )
    cysvm_grouped[key].append(record)

severity_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}

sorted_cysvm = sorted(
    cysvm_grouped.items(),
    key=lambda item: (
        severity_order.get(item[0][1], 99),
        item[0][2],
        item[0][3],
        item[0][0],
    ),
)

for index, ((title, severity, host, port), records) in enumerate(sorted_cysvm, 1):
    targets = sorted({item["URL / Target"] for item in records if item["URL / Target"]})
    evidences = sorted({item["Evidence"] for item in records if item["Evidence"]})

    cysvm_rows.append(
        {
            "No": index,
            "Nama Temuan": title,
            "Severity Risk": severity,
            "URL": safe_join(targets),
            "Keterangan": safe_join(evidences),
            "Rekomendasi": "",
        }
    )

cysvm_df = pd.DataFrame(
    cysvm_rows,
    columns=[
        "No",
        "Nama Temuan",
        "Severity Risk",
        "URL",
        "Keterangan",
        "Rekomendasi",
    ],
)


# ============================================================
# WRITE EXCEL
# ============================================================

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    results_df.to_excel(writer, index=False, sheet_name="Results")
    remediation_df.to_excel(writer, index=False, sheet_name="Remediasi")
    cysvm_df.to_excel(writer, index=False, sheet_name="CYSVM")


# ============================================================
# FORMAT EXCEL
# ============================================================

workbook = load_workbook(OUTPUT_FILE)
expected_sheets = ["Results", "Remediasi", "CYSVM"]

for sheet_name in list(workbook.sheetnames):
    if sheet_name not in expected_sheets:
        del workbook[sheet_name]

workbook._sheets = [workbook[sheet_name] for sheet_name in expected_sheets]

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)
column_widths = {
    "Results": {
        "A": 6,
        "B": 42,
        "C": 10,
        "D": 45,
        "E": 12,
        "F": 100,
        "G": 100,
    },
    "Remediasi": {
        "A": 6,
        "B": 34,
        "C": 70,
        "D": 95,
        "E": 16,
    },
    "CYSVM": {
        "A": 6,
        "B": 45,
        "C": 16,
        "D": 90,
        "E": 70,
        "F": 95,
    },
}

for worksheet in workbook.worksheets:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_letter, width in column_widths.get(worksheet.title, {}).items():
        worksheet.column_dimensions[column_letter].width = width

workbook.save(OUTPUT_FILE)


# ============================================================
# SANITY CHECK + SUMMARY PRINT
# ============================================================

check_workbook = load_workbook(OUTPUT_FILE, read_only=True, data_only=True)
results_ws = check_workbook["Results"]
cysvm_ws = check_workbook["CYSVM"]

results_headers = [
    results_ws.cell(row=1, column=column_index).value
    for column_index in range(1, results_ws.max_column + 1)
]

cysvm_headers = [
    cysvm_ws.cell(row=1, column=column_index).value
    for column_index in range(1, cysvm_ws.max_column + 1)
]

summary = {
    "Input Files": [OPEN_PORTS_FILE, FFUF_FILE] + NUCLEI_FILES,
    "Total Open Port Rows": len(ports),
    "Total Unique FFUF Findings": len(ffuf_seen),
    "Total Unique Nuclei Findings": len(nuclei_records),
    "Total Results Rows": len(results_rows),
    "Total CYSVM Rows": len(cysvm_rows),
    "Unmatched FFUF HostPort Not Added To Results": len(unmatched_ffuf_keys),
    "Unmatched Nuclei HostPort Not Added To Results": len(unmatched_nuclei_keys),
    "Sheets": check_workbook.sheetnames,
    "Sheets OK": check_workbook.sheetnames == expected_sheets,
    "Results Columns OK": results_headers == [
        "No",
        "Host / IP Address",
        "Port",
        "Service",
        "Protocol",
        "FFUF URLs",
        "Nuclei Findings",
    ],
    "CYSVM Columns OK": cysvm_headers == [
        "No",
        "Nama Temuan",
        "Severity Risk",
        "URL",
        "Keterangan",
        "Rekomendasi",
    ],
    "Output": OUTPUT_FILE,
}

print(json.dumps(summary, indent=2, ensure_ascii=False))
